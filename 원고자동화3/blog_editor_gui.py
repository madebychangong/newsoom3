#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
블로그 원고 자동 수정 프로그램 (GUI 버전)
Claude API 기반
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import openpyxl
import anthropic
import re
import os
from datetime import datetime
import threading
import json
import base64

class BlogEditorGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("블로그 원고 자동 수정 프로그램")
        self.root.geometry("800x700")
        self.root.resizable(False, False)
        
        # 데이터 저장 변수
        self.api_key = ""
        self.selected_model = "claude-sonnet-4-5"  # 기본값
        self.forbidden_words = {}
        self.examples = []
        self.input_file = ""
        self.is_processing = False
        self.config_file = os.path.join(os.path.expanduser("~"), ".blog_editor_config.json")

        self.setup_ui()
        self.load_saved_api_key()  # 저장된 API 키 불러오기
        
    def load_saved_api_key(self):
        """저장된 API 키 불러오기"""
        try:
            if os.path.exists(self.config_file):
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    saved_key = config.get('api_key', '')
                    
                    if saved_key:
                        # 간단한 디코딩 (보안은 약하지만 평문보다는 나음)
                        decoded_key = base64.b64decode(saved_key).decode('utf-8')
                        self.api_key = decoded_key
                        self.api_entry.insert(0, decoded_key)
                        self.api_status.config(text="✅ 저장된 API 키 불러옴", fg="green")
                        self.log("✅ 저장된 API 키를 불러왔습니다", "#27ae60")
                        self.check_ready()
        except Exception as e:
            self.log(f"⚠️  저장된 API 키 불러오기 실패: {str(e)}", "#e67e22")
    
    def save_api_key_to_file(self):
        """API 키를 파일로 저장"""
        try:
            # 간단한 인코딩
            encoded_key = base64.b64encode(self.api_key.encode('utf-8')).decode('utf-8')
            
            config = {'api_key': encoded_key}
            
            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(config, f)
                
        except Exception as e:
            self.log(f"⚠️  API 키 저장 실패: {str(e)}", "#e67e22")
        
    def setup_ui(self):
        """UI 구성"""
        # 메인 프레임
        main_frame = ttk.Frame(self.root, padding="15")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 제목
        title_frame = ttk.Frame(main_frame)
        title_frame.pack(fill=tk.X, pady=(0, 20))
        
        title_label = tk.Label(title_frame, text="📝 블로그 원고 자동 수정 프로그램",
                              font=("맑은 고딕", 18, "bold"), fg="#2c3e50")
        title_label.pack()

        subtitle_label = tk.Label(title_frame, text="Claude AI 기반",
                                 font=("맑은 고딕", 10), fg="#7f8c8d")
        subtitle_label.pack()
        
        # 1. API 키 및 모델 선택 섹션
        api_frame = ttk.LabelFrame(main_frame, text="  1️⃣  Claude API 설정  ", padding="10")
        api_frame.pack(fill=tk.X, pady=(0, 10))

        # API 키 입력
        api_input_frame = ttk.Frame(api_frame)
        api_input_frame.pack(fill=tk.X)

        self.api_entry = ttk.Entry(api_input_frame, width=50, show="*", font=("맑은 고딕", 10))
        self.api_entry.pack(side=tk.LEFT, padx=(0, 10), fill=tk.X, expand=True)

        self.api_button = ttk.Button(api_input_frame, text="저장", command=self.save_api_key)
        self.api_button.pack(side=tk.LEFT)

        self.api_status = tk.Label(api_frame, text="❌ API 키 미등록",
                                   font=("맑은 고딕", 9), fg="red")
        self.api_status.pack(anchor=tk.W, pady=(5, 0))

        api_help = tk.Label(api_frame, text="💡 API 키 발급: https://console.anthropic.com/",
                           font=("맑은 고딕", 8), fg="#3498db", cursor="hand2")
        api_help.pack(anchor=tk.W)
        api_help.bind("<Button-1>", lambda e: self.open_url("https://console.anthropic.com/"))

        # 모델 선택
        model_frame = ttk.Frame(api_frame)
        model_frame.pack(fill=tk.X, pady=(10, 0))

        tk.Label(model_frame, text="모델 선택:", font=("맑은 고딕", 9)).pack(side=tk.LEFT, padx=(0, 10))

        self.model_var = tk.StringVar(value="claude-sonnet-4-5")
        model_combo = ttk.Combobox(model_frame, textvariable=self.model_var,
                                   values=[
                                       "claude-sonnet-4-5",
                                       "claude-haiku-4-5"
                                   ], state="readonly", width=30)
        model_combo.pack(side=tk.LEFT)
        model_combo.bind("<<ComboboxSelected>>", self.on_model_change)

        self.model_info = tk.Label(model_frame, text="(규칙 준수: 최고, 비용: ₩26/개)",
                                   font=("맑은 고딕", 8), fg="#7f8c8d")
        self.model_info.pack(side=tk.LEFT, padx=(10, 0))
        
        # 2. 파일 선택 섹션
        file_frame = ttk.LabelFrame(main_frame, text="  2️⃣  수정할 엑셀 파일 선택  ", padding="10")
        file_frame.pack(fill=tk.X, pady=(0, 10))
        
        file_input_frame = ttk.Frame(file_frame)
        file_input_frame.pack(fill=tk.X)
        
        self.file_label = tk.Label(file_input_frame, text="선택된 파일 없음", 
                                   font=("맑은 고딕", 9), fg="gray", anchor=tk.W)
        self.file_label.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        self.file_button = ttk.Button(file_input_frame, text="📁 파일 선택", command=self.select_file)
        self.file_button.pack(side=tk.LEFT)
        
        file_help = tk.Label(file_frame, text="💡 같은 폴더에 금칙어_리스트.xlsx 필요",
                            font=("맑은 고딕", 8), fg="#7f8c8d")
        file_help.pack(anchor=tk.W, pady=(5, 0))
        
        # 3. 실행 버튼
        run_frame = ttk.LabelFrame(main_frame, text="  3️⃣  자동 수정 실행  ", padding="10")
        run_frame.pack(fill=tk.X, pady=(0, 10))
        
        self.run_button = ttk.Button(run_frame, text="🚀 자동 수정 시작", 
                                     command=self.start_processing, state='disabled')
        self.run_button.pack(fill=tk.X)
        
        # 4. 진행 상황
        progress_frame = ttk.LabelFrame(main_frame, text="  처리 상황  ", padding="10")
        progress_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        self.progress_text = scrolledtext.ScrolledText(progress_frame, width=80, height=20, 
                                                       wrap=tk.WORD, font=("맑은 고딕", 9))
        self.progress_text.pack(fill=tk.BOTH, expand=True)
        
        # 하단 상태바
        status_frame = ttk.Frame(main_frame)
        status_frame.pack(fill=tk.X)
        
        self.status_label = tk.Label(status_frame, text="대기 중...", 
                                     font=("맑은 고딕", 9), fg="#3498db", anchor=tk.W)
        self.status_label.pack(side=tk.LEFT)
        
    def open_url(self, url):
        """URL 열기"""
        import webbrowser
        webbrowser.open(url)

    def on_model_change(self, event=None):
        """모델 선택 변경 시 처리"""
        selected = self.model_var.get()
        self.selected_model = selected

        if selected == "claude-sonnet-4-5":
            self.model_info.config(text="(규칙 준수: 최고, 비용: ₩26/개)")
            self.log("✅ 모델 변경: Claude Sonnet 4.5 (최고 정확도)", "#3498db")
        elif selected == "claude-haiku-4-5":
            self.model_info.config(text="(규칙 준수: 높음, 비용: ₩9/개)")
            self.log("✅ 모델 변경: Claude Haiku 4.5 (빠른 처리)", "#3498db")
        
    def log(self, message, color=None):
        """로그 출력"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.progress_text.insert(tk.END, f"[{timestamp}] {message}\n")
        if color:
            # 마지막 줄에 색상 적용
            line_start = self.progress_text.index("end-2c linestart")
            line_end = self.progress_text.index("end-1c")
            self.progress_text.tag_add(color, line_start, line_end)
            self.progress_text.tag_config(color, foreground=color)
        self.progress_text.see(tk.END)
        self.root.update()
        
    def save_api_key(self):
        """API 키 저장 (검증 없이)"""
        self.api_key = self.api_entry.get().strip()
        
        if not self.api_key:
            messagebox.showwarning("경고", "API 키를 입력해주세요.")
            return
        
        # 파일로 저장
        self.save_api_key_to_file()

        self.api_status.config(text=f"✅ API 키 저장 완료 ({self.model_var.get()})", fg="green")
        self.log("✅ API 키가 저장되었습니다", "#27ae60")
        self.check_ready()
        messagebox.showinfo("저장 완료", "API 키가 저장되었습니다.\n다음 실행 시 자동으로 불러옵니다.")
            
    def select_file(self):
        """파일 선택"""
        filename = filedialog.askopenfilename(
            title="수정할 엑셀 파일 선택",
            filetypes=[("Excel 파일", "*.xlsx"), ("모든 파일", "*.*")]
        )
        
        if filename:
            self.input_file = filename
            basename = os.path.basename(filename)
            self.file_label.config(text=f"📄 {basename}", fg="green")
            self.log(f"파일 선택: {basename}", "#3498db")
            self.check_ready()
            
    def check_ready(self):
        """실행 가능 여부 체크"""
        if self.api_key and self.input_file:
            self.run_button.config(state='normal')
            self.status_label.config(text="✅ 준비 완료 - 실행 버튼을 눌러주세요", fg="green")
        else:
            self.run_button.config(state='disabled')
            
    def start_processing(self):
        """처리 시작 (별도 스레드)"""
        if self.is_processing:
            return
        
        self.is_processing = True
        self.run_button.config(state='disabled')
        self.file_button.config(state='disabled')
        self.api_button.config(state='disabled')
        
        # 별도 스레드에서 처리
        thread = threading.Thread(target=self.process_file)
        thread.daemon = True
        thread.start()
        
    def load_forbidden_words(self, base_dir):
        """금칙어 로딩"""
        try:
            file_path = os.path.join(base_dir, '금칙어_리스트.xlsx')
            
            if not os.path.exists(file_path):
                self.log(f"⚠️  금칙어 파일 없음: {file_path}", "#e67e22")
                return False
            
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            for row_idx in range(3, ws.max_row + 1):
                forbidden = ws.cell(row_idx, 2).value
                alternatives = []
                
                for col_idx in range(3, 10):
                    alt = ws.cell(row_idx, col_idx).value
                    if alt:
                        alternatives.append(str(alt).strip())
                
                if forbidden and alternatives:
                    self.forbidden_words[str(forbidden).strip()] = alternatives
            
            self.log(f"✅ 금칙어 {len(self.forbidden_words)}개 로딩 완료", "#27ae60")
            return True
            
        except Exception as e:
            self.log(f"❌ 금칙어 로딩 실패: {str(e)}", "#e74c3c")
            return False
            
    def load_examples(self, base_dir):
        """학습 예시 로딩"""
        try:
            # 수정전후.xlsx
            file1 = os.path.join(base_dir, '수정전후.xlsx')
            if os.path.exists(file1):
                wb1 = openpyxl.load_workbook(file1)
                ws1 = wb1.active
                
                for row_idx in range(2, ws1.max_row + 1):
                    example = {
                        'keyword': ws1.cell(row_idx, 2).value,
                        'char_count': ws1.cell(row_idx, 3).value,
                        'main_keyword_count': ws1.cell(row_idx, 4).value,
                        'sub_keyword_count': ws1.cell(row_idx, 5).value,
                        'extra_keyword_count': ws1.cell(row_idx, 6).value,
                        'original': ws1.cell(row_idx, 7).value,
                        'edited': ws1.cell(row_idx, 8).value
                    }
                    if example['original'] and example['edited']:
                        self.examples.append(example)
            
            # 블로그_작업_엑셀템플릿.xlsx
            file2 = os.path.join(base_dir, '블로그_작업_엑셀템플릿.xlsx')
            if os.path.exists(file2):
                wb2 = openpyxl.load_workbook(file2)
                ws2 = wb2.active
                
                for row_idx in range(2, ws2.max_row + 1):
                    example = {
                        'keyword': ws2.cell(row_idx, 2).value,
                        'char_count': ws2.cell(row_idx, 3).value,
                        'main_keyword_count': ws2.cell(row_idx, 4).value,
                        'sub_keyword_count': ws2.cell(row_idx, 5).value,
                        'extra_keyword_count': ws2.cell(row_idx, 6).value,
                        'original': ws2.cell(row_idx, 7).value,
                        'edited': ws2.cell(row_idx, 8).value
                    }
                    if example['original'] and example['edited']:
                        self.examples.append(example)
            
            self.log(f"✅ 학습 예시 {len(self.examples)}개 로딩 완료", "#27ae60")
            return len(self.examples) > 0
            
        except Exception as e:
            self.log(f"❌ 예시 로딩 실패: {str(e)}", "#e74c3c")
            return False
            
    def analyze_speaker(self, text, client):
        """화자 정보 분석 (성별, 연령대, 상황)"""
        if not text:
            return "분석 불가"

        try:
            analysis_prompt = f"""
다음 블로그 글을 분석하여 작성자(화자)의 정보를 유추해주세요.

글:
{text[:500]}...

다음 형식으로만 답변하세요 (다른 설명 없이):
성별: [남성/여성/알 수 없음]
연령대: [20대/30대/40대/50대/60대 이상/알 수 없음]
상황: [한 줄로 간단히 설명]

예시:
성별: 여성
연령대: 30대
상황: 자녀 키 성장 고민
"""

            message = client.messages.create(
                model=self.selected_model,
                max_tokens=1024,
                messages=[
                    {"role": "user", "content": analysis_prompt}
                ]
            )
            analysis = message.content[0].text.strip()

            # 한 줄로 정리
            analysis = analysis.replace('\n', ' / ')

            return analysis

        except Exception as e:
            return f"분석 실패: {str(e)}"
    
    def add_line_breaks(self, text):
        """문장마다 줄바꿈 추가"""
        if not text:
            return text

        # 문장 종결 부호 뒤에 줄바꿈 추가
        # 이미 줄바꿈이 있으면 추가하지 않음
        text = re.sub(r'([.!?])\s+', r'\1\n', text)

        # 연속된 줄바꿈을 2개로 제한 (문단 구분 보존)
        text = re.sub(r'\n{3,}', '\n\n', text)

        return text.strip()
    
    def apply_basic_corrections(self, text):
        """기본 교정"""
        if not text:
            return text
        
        # 1. 네요 -> 내요 (무조건)
        text = text.replace('네요', '내요')
        
        # 2. 더라 -> 더 라 (무조건)
        text = text.replace('더라', '더 라')
        
        # 3. 이모티콘 앞뒤 띄어쓰기 처리 (서브키워드 카운팅을 위해)
        emoticons = ['^^', '??', '!!', '~~', '...', 'ㅠㅠ', 'TT', 'ㅎㅎ', ';;', '--', 'ㅋㅋ']
        
        for emoticon in emoticons:
            # 이모티콘 앞에 띄어쓰기 없으면 추가
            # "좋아요^^" → "좋아요 ^^"
            text = re.sub(r'([^\s])' + re.escape(emoticon), r'\1 ' + emoticon, text)
            
            # 이모티콘 뒤 문장부호 제거하고 띄어쓰기
            # "^^ ." → "^^ "
            text = text.replace(f'{emoticon}.', f'{emoticon} ')
            text = text.replace(f'{emoticon},', f'{emoticon} ')
            text = text.replace(f'{emoticon}!', f'{emoticon} ')
            text = text.replace(f'{emoticon}?', f'{emoticon} ')
            
            # 이모티콘 뒤에 아무것도 없거나 문자가 바로 오면 띄어쓰기 추가
            # "^^ 다음" 은 그대로, "^^다음" → "^^ 다음"
            text = re.sub(re.escape(emoticon) + r'([^\s.,!?])', emoticon + r' \1', text)

        # 금칙어 치환은 AI가 직접 하도록 제거 (테스트용)

        return text
    
    def clean_markdown(self, text):
        """마크다운 형식 제거"""
        if not text:
            return text
        
        # ** 강조 제거
        text = re.sub(r'\*\*([^*]+)\*\*', r'\1', text)
        
        # * 강조 제거
        text = re.sub(r'\*([^*]+)\*', r'\1', text)
        
        # # 헤더 제거
        text = re.sub(r'^#+\s+', '', text, flags=re.MULTILINE)
        
        # 마크다운 코드 블록 제거 (```로 둘러싸인 부분)
        text = re.sub(r'```[^`]*```', '', text, flags=re.DOTALL)
        
        return text.strip()
    
    def parse_keyword_rule(self, rule_text):
        """키워드 규칙 파싱"""
        if not rule_text:
            return ""
        
        rule_text = str(rule_text).strip()
        
        # "키워드 : 숫자" 형식 파싱
        match = re.match(r'(.+?)\s*:\s*(\d+)', rule_text)
        if match:
            keyword = match.group(1).strip()
            count = match.group(2).strip()
            return f"'{keyword}'를 정확히 {count}번 반복 (±1 허용)"
        
        return rule_text
    
    def parse_sub_keywords(self, rule_text):
        """조각 키워드 규칙 파싱"""
        if not rule_text:
            return ""
        
        rule_text = str(rule_text).strip()
        
        # 여러 줄로 나뉜 경우 처리
        lines = rule_text.split('\n')
        parsed_rules = []
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
            
            # "키워드 : 숫자" 형식 파싱
            match = re.match(r'(.+?)\s*:\s*(\d+)', line)
            if match:
                keyword = match.group(1).strip()
                count = match.group(2).strip()
                parsed_rules.append(f"'{keyword}' {count}번")
        
        if parsed_rules:
            return ", ".join(parsed_rules) + " 각각 반복 (±1 허용)"
        
        return rule_text
        
    def create_system_prompt(self):
        """시스템 프롬프트 생성 (고정 부분 - 캐싱용)"""

        # 금칙어 리스트 생성
        forbidden_list = ""
        for forbidden, alternatives in self.forbidden_words.items():
            alt_text = ", ".join(alternatives[:3])  # 최대 3개까지만
            forbidden_list += f"- '{forbidden}' 대신 → {alt_text} 중 문맥에 맞는 것 사용\n"

        system_prompt = f"""당신은 원고를 정확한 규칙에 맞춰 수정하는 전문가입니다.

# 필수 규칙 (모두 100% 준수 필수!)

## 1. 글자수 (최우선 준수!)
- 목표 글자수의 ±5% 범위 내로 작성
- **이 규칙을 반드시 지켜야 합니다**
- 글자수 초과 시: 불필요한 형용사, 부사, 중복 표현 삭제
- 글자수 부족 시: 구체적인 예시, 부연 설명, 경험담 추가
- 작성 후 반드시 글자수를 세어서 범위 내인지 확인

## 2. 키워드 규칙
- **통 키워드 (핵심 키워드)**: 지정된 횟수만큼 반복
  → **중요**: 이 횟수는 첫 문단을 제외한 나머지 문단에서의 반복 횟수
  → 첫 문단에는 무조건 2회, 나머지 문단에서만 지정된 횟수 반복
- **조각 키워드**: 지정된 횟수만큼 반복
  → **중요**: 이 횟수도 첫 문단을 제외한 나머지 문단에서의 반복 횟수
- **서브 키워드 목록 수**: 지정된 개수 달성
  → 조각 키워드를 제외한 2회 이상 등장하는 단어의 총 개수
  → **중요**: 단어가 부족하면 중복 문장부호 적극 활용 (^^, ??, !!, ~~, .., ㅠㅠ, TT, ㅎㅎ 등)
  → 각 중복 문장부호는 서브키워드 1개로 카운팅됨
  → 예시: ^^ 사용, ?? 사용, .. 사용, ~~ 사용 등으로 자연스럽게 개수 채우기

## 3. 카운팅 규칙 (매우 중요!)
- **띄어쓰기 단위로 카운팅**
- **모든 키워드 뒤에는 반드시 띄어쓰기 필수!**
  → "강남 맛집 추천을" (X - 조사 붙음, 카운팅 안됨)
  → "강남 맛집 추천 리스트" (O - 띄어쓰기 유지, 카운팅 됨)
  → "강남 맛집 추천 정보를" (O - 띄어쓰기 있음, 카운팅 됨)
- **한글자 조사(을/를/이/가)는 절대 키워드에 붙이지 말 것**
  → 키워드 뒤 한글자 조사가 필요하면 띄어쓰기 + 우회 표현 사용
  → 예: "강남 맛집 추천을" (X) → "강남 맛집 추천 정보를" (O)
- **두글자 이상 조사(으로/에게/부터)**: 띄어쓰기 허용
- **중복 문장부호 카운팅**: 앞뒤 띄어쓰기 필수
  → "궁금해요 ^^ 정말" (O - ^^ 는 1개 서브키워드)
  → "그렇내요.." (X - 카운팅 안됨, 띄어쓰기 없음)
  → "그렇내요 .." (O - 카운팅 됨, 띄어쓰기 있음)

## 4. 첫 문단 필수 규칙 (매우 중요!)
- **첫 문단은 무조건 4문장 이상으로 작성**
- **첫 문단에 핵심 키워드 정확히 2회 등장 필수**
- 핵심 키워드 사이에 2문장 이상 삽입
- 예시: "페퍼로니피자 다이어트 관련해서 요즘 알아보고 있어요. (중간 2문장 이상) 페퍼로니피자 다이어트 정보를 찾아보니... (추가 문장들로 4문장 이상 채우기)"
- **주의**: 첫 문단은 첫 번째 문단 구분(줄바꿈) 전까지를 의미함

## 5. 핵심 키워드로 시작하는 문장
- 글 전체에서 핵심 키워드로 시작하는 문장이 지정된 개수만큼 있어야 함
- 예: "강남 맛집 추천을 받아서..." (X - 조사 붙음)
- 예: "강남 맛집 추천 리스트를 보면..." (O - 띄어쓰기 유지)

## 6. 글 구조 및 문단 구분
- **도입부**: 고민/궁금증/경험 소개
- **중간부**: 자연스러운 키워드 반복
- **마무리**: 댓글 유도 (정보 공유 요청, 질문 등)
- **문단 구분**: 2~4문장마다 적절히 빈 줄(줄바꿈 2번)로 문단을 나누어 가독성 높이기

## 7. 키워드 부족 시
- **일반 단어 부족**: 자연스러운 문맥에 추가 삽입
- **서브키워드 부족**: 중복 문장부호를 적극 활용하여 채우기
  → ^^, ??, !!, ~~, ..., ㅠㅠ, TT, ㅎㅎ 등을 문장 끝이나 중간에 자연스럽게 배치
  → 각 중복 문장부호는 앞뒤 띄어쓰기 필수 (예: "궁금해요 ^^ 정말" / "그렇네요 ...")
  → 개수가 다르면 다른 서브키워드 (예: ?? 와 ??? 는 별개)
- **그래도 부족하면**: 마지막에 #해시태그 형식으로 추가
  → 맛집 서브키워드 추가시 예: # 강남 맛집 # 맛집 추천

## 8. 금칙어 (절대 사용 금지)
**다음 단어들은 절대 사용하지 말고, 문맥에 맞는 대체어를 사용하세요:**

{forbidden_list}

**⚠️ 모든 규칙은 동등하게 중요! 하나라도 어기면 안 됨!**

**수정된 원고만 출력**하고, 설명이나 주석은 절대 붙이지 마세요."""

        return system_prompt

    def create_user_prompt(self, row_data):
        """유저 프롬프트 생성 (변동 부분)"""

        # 키워드 규칙 파싱
        main_keyword_rule = self.parse_keyword_rule(row_data['main_keyword_count'])
        sub_keyword_rule = self.parse_sub_keywords(row_data['sub_keyword_count'])
        extra_keyword_count = str(row_data['extra_keyword_count']).strip() if row_data['extra_keyword_count'] else "0"

        # 글자수 및 오차 계산
        target_chars = int(row_data['char_count']) if row_data['char_count'] else 1000
        char_tolerance = int(target_chars * 0.05)  # 5% 오차

        # 통키워드 문장 시작 횟수
        keyword_start_count = str(row_data['keyword_start_count']).strip() if row_data['keyword_start_count'] else "2~3"

        # 원고에서 제목 라인 제거 (맨 위 # 으로 시작하는 한 줄)
        original_text = row_data['original']
        if original_text.strip().startswith('#'):
            # 첫 줄 제거
            lines = original_text.split('\n', 1)
            original_text = lines[1] if len(lines) > 1 else ""
            original_text = original_text.strip()

        user_prompt = f"""# 수정 조건

**키워드**: {row_data['keyword']}
**통 키워드**: {main_keyword_rule}
**조각 키워드**: {sub_keyword_rule}
**서브 키워드 목록 수**: {extra_keyword_count}개
**목표 글자수**: {target_chars}자 (허용 범위: {target_chars - char_tolerance}~{target_chars + char_tolerance}자)
**통키워드로 시작하는 문장**: {keyword_start_count}개

# 수정할 원고

{original_text}

# 검수 체크리스트
- [ ] 글자수: {target_chars - char_tolerance}~{target_chars + char_tolerance}자 범위 내
- [ ] 첫 문단: 4문장 이상 + '{row_data['keyword']}' 정확히 2회
- [ ] 통키워드/조각키워드: 지정 횟수 준수
- [ ] 서브키워드: {extra_keyword_count}개 달성
- [ ] 통키워드로 시작하는 문장: {keyword_start_count}개
- [ ] 금칙어: 0개 (전부 대체어 사용)
- [ ] 문단 구분: 2~4문장마다 빈 줄"""

        return user_prompt
        
    def process_file(self):
        """파일 처리 메인 로직"""
        try:
            self.log("\n" + "="*60, "#2c3e50")
            self.log("🚀 자동 수정 시작...", "#2c3e50")
            self.log("="*60, "#2c3e50")
            self.status_label.config(text="⏳ 처리 중...", fg="orange")
            
            # 같은 폴더 경로
            base_dir = os.path.dirname(self.input_file)
            
            # 금칙어 로딩
            if not self.load_forbidden_words(base_dir):
                messagebox.showwarning("경고", "금칙어 파일을 찾을 수 없습니다.\n같은 폴더에 '금칙어_리스트.xlsx'를 넣어주세요.")
            
            # 입력 파일 로드
            wb = openpyxl.load_workbook(self.input_file)
            ws = wb.active

            # Claude 클라이언트 초기화
            client = anthropic.Anthropic(api_key=self.api_key)

            total_rows = ws.max_row - 1
            
            for row_idx in range(2, ws.max_row + 1):
                self.log(f"\n{'─'*60}", "#95a5a6")
                self.log(f"📄 {row_idx-1}/{total_rows}번째 원고 처리 중...", "#3498db")
                self.log(f"{'─'*60}", "#95a5a6")
                
                # 데이터 추출 (새 양식)
                row_data = {
                    'keyword': ws.cell(row_idx, 2).value,  # B열: 키워드
                    'main_keyword_count': ws.cell(row_idx, 4).value,  # D열: 통키워드 반복수
                    'sub_keyword_count': ws.cell(row_idx, 5).value,  # E열: 조각키워드 반복수
                    'original': ws.cell(row_idx, 7).value,  # G열: 원고
                    'char_count': ws.cell(row_idx, 10).value,  # J열: 실제 글자수
                    'keyword_start_count': ws.cell(row_idx, 11).value,  # K열: 문장시작통키워드 수
                    'extra_keyword_count': ws.cell(row_idx, 12).value  # L열: 보정 서브키워드 목록 수
                }
                
                if not row_data['original']:
                    self.log(f"⚠️  {row_idx}행: 원고 없음, 건너뜀", "#e67e22")
                    continue
                
                self.log(f"키워드: {row_data['keyword']}")
                self.log(f"목표 글자수: {row_data['char_count']}자")
                
                # AI 수정
                self.log("⏳ AI 수정 중... (10~30초 소요, 캐싱 적용)", "#f39c12")
                system_prompt = self.create_system_prompt()
                user_prompt = self.create_user_prompt(row_data)

                message = client.messages.create(
                    model=self.selected_model,
                    max_tokens=4096,
                    system=[
                        {
                            "type": "text",
                            "text": system_prompt,
                            "cache_control": {"type": "ephemeral"}
                        }
                    ],
                    messages=[
                        {"role": "user", "content": user_prompt}
                    ]
                )
                edited_text = message.content[0].text.strip()
                
                # 마크다운 형식 제거
                edited_text = self.clean_markdown(edited_text)
                
                # AI 생성 후 기본 교정 적용 (네요→내요, 더라→더 라, 금칙어)
                edited_text = self.apply_basic_corrections(edited_text)
                
                # 문장마다 줄바꿈 추가
                edited_text = self.add_line_breaks(edited_text)
                
                # 결과 저장 (M열 = 13번)
                ws.cell(row_idx, 13).value = edited_text
                self.log(f"✅ AI 수정 및 교정 완료 (결과 글자수: {len(edited_text)}자)", "#27ae60")
                
                # 화자 분석 (N열 = 14번)
                self.log("⏳ 화자 정보 분석 중...", "#3498db")
                speaker_info = self.analyze_speaker(edited_text, client)
                ws.cell(row_idx, 14).value = speaker_info
                self.log(f"✅ 화자 분석 완료: {speaker_info}", "#27ae60")
                
            # 결과 파일 저장 (원본 파일에 덮어쓰기)
            wb.save(self.input_file)
            
            self.log("\n" + "="*60, "#2c3e50")
            self.log("🎉 모든 작업 완료!", "#27ae60")
            self.log("="*60, "#2c3e50")
            self.log(f"📁 저장 위치: {self.input_file}", "#3498db")
            
            self.status_label.config(text="✅ 완료!", fg="green")
            
            messagebox.showinfo("완료", f"수정이 완료되었습니다!\n\n원본 파일에 저장됨:\n{self.input_file}")
            
        except Exception as e:
            self.log(f"\n❌ 오류 발생: {str(e)}", "#e74c3c")
            self.status_label.config(text="❌ 오류 발생", fg="red")
            messagebox.showerror("오류", f"처리 중 오류가 발생했습니다:\n{str(e)}")
            
        finally:
            self.is_processing = False
            self.run_button.config(state='normal')
            self.file_button.config(state='normal')
            self.api_button.config(state='normal')

def main():
    root = tk.Tk()
    app = BlogEditorGUI(root)
    root.mainloop()

if __name__ == "__main__":
    main()
