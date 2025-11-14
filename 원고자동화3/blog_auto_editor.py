#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
블로그 원고 자동 수정 프로그램
Gemini API 기반
"""

import openpyxl
import google.generativeai as genai
import re
import os
from datetime import datetime

class BlogEditor:
    def __init__(self):
        self.api_key = ""
        self.forbidden_words = {}
        self.examples = []
        
        print("="*60)
        print("📝 블로그 원고 자동 수정 프로그램")
        print("="*60)
        
    def log(self, message):
        """로그 출력"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        print(f"[{timestamp}] {message}")
        
    def load_forbidden_words(self):
        """금칙어 리스트 로딩"""
        try:
            wb = openpyxl.load_workbook('/mnt/user-data/uploads/금칙어_리스트.xlsx')
            ws = wb.active
            
            for row_idx in range(3, ws.max_row + 1):
                forbidden = ws.cell(row_idx, 2).value
                alternatives = []
                
                for col_idx in range(3, 10):
                    alt = ws.cell(row_idx, col_idx).value
                    if alt:
                        alternatives.append(str(alt).strip())
                
                if forbidden and alternatives:
                    self.forbidden_words[str(forbidden).strip()] = alternatives
            
            self.log(f"✅ 금칙어 {len(self.forbidden_words)}개 로딩 완료")
            return True
            
        except Exception as e:
            self.log(f"❌ 금칙어 로딩 실패: {str(e)}")
            return False
            
    def load_examples(self):
        """학습 예시 로딩"""
        try:
            # 수정전후.xlsx
            wb1 = openpyxl.load_workbook('/mnt/user-data/uploads/수정전후.xlsx')
            ws1 = wb1.active
            
            for row_idx in range(2, ws1.max_row + 1):
                example = {
                    'keyword': ws1.cell(row_idx, 2).value,
                    'char_count': ws1.cell(row_idx, 3).value,
                    'main_keyword_count': ws1.cell(row_idx, 4).value,
                    'sub_keyword_count': ws1.cell(row_idx, 5).value,
                    'extra_keyword_count': ws1.cell(row_idx, 6).value,
                    'original': ws1.cell(row_idx, 7).value,
                    'edited': ws1.cell(row_idx, 8).value
                }
                if example['original'] and example['edited']:
                    self.examples.append(example)
            
            # 블로그_작업_엑셀템플릿.xlsx
            wb2 = openpyxl.load_workbook('/mnt/user-data/uploads/블로그_작업_엑셀템플릿.xlsx')
            ws2 = wb2.active
            
            for row_idx in range(2, ws2.max_row + 1):
                example = {
                    'keyword': ws2.cell(row_idx, 2).value,
                    'char_count': ws2.cell(row_idx, 3).value,
                    'main_keyword_count': ws2.cell(row_idx, 4).value,
                    'sub_keyword_count': ws2.cell(row_idx, 5).value,
                    'extra_keyword_count': ws2.cell(row_idx, 6).value,
                    'original': ws2.cell(row_idx, 7).value,
                    'edited': ws2.cell(row_idx, 8).value
                }
                if example['original'] and example['edited']:
                    self.examples.append(example)
            
            self.log(f"✅ 학습 예시 {len(self.examples)}개 로딩 완료")
            return True
            
        except Exception as e:
            self.log(f"❌ 예시 로딩 실패: {str(e)}")
            return False
            
    def input_api_key(self):
        """API 키 입력"""
        print("\n" + "="*60)
        print("1️⃣  Gemini API 키를 입력해주세요")
        print("=" *60)
        print("💡 API 키 발급: https://aistudio.google.com/app/apikey")
        
        self.api_key = input("\n🔑 API 키: ").strip()
        
        if not self.api_key:
            print("❌ API 키가 입력되지 않았습니다.")
            return False
        
        # API 키 검증
        try:
            genai.configure(api_key=self.api_key)
            model = genai.GenerativeModel('gemini-2.5-pro')
            response = model.generate_content("안녕")
            
            self.log("✅ Gemini API 연결 성공! (모델: gemini-2.5-pro)")
            return True
            
        except Exception as e:
            print(f"❌ API 키 검증 실패: {str(e)}")
            return False
            
    def select_file(self):
        """파일 선택"""
        print("\n" + "="*60)
        print("2️⃣  수정할 엑셀 파일 경로를 입력해주세요")
        print("="*60)
        
        file_path = input("\n📁 파일 경로: ").strip()
        
        if not os.path.exists(file_path):
            print(f"❌ 파일을 찾을 수 없습니다: {file_path}")
            return None
        
        if not file_path.endswith('.xlsx'):
            print("❌ .xlsx 파일만 지원됩니다.")
            return None
        
        self.log(f"✅ 파일 선택: {os.path.basename(file_path)}")
        return file_path
        
    def apply_basic_corrections(self, text):
        """기본 교정"""
        if not text:
            return text
        
        # 1. 네요 -> 내요
        text = re.sub(r'네요', '내요', text)
        
        # 2. ~더라 -> ~더 라
        text = re.sub(r'(\w)더라(\s|$|[,.])', r'\1더 라\2', text)
        text = re.sub(r'더라구요', '더 라구요', text)
        
        # 3. 금칙어 치환
        for forbidden, alternatives in self.forbidden_words.items():
            if forbidden in text and alternatives:
                text = text.replace(forbidden, alternatives[0])
        
        return text
        
    def create_prompt(self, row_data):
        """Gemini용 프롬프트 생성"""
        
        # 예시 데이터 (처음 3개)
        examples_text = ""
        for i, ex in enumerate(self.examples[:3], 1):
            examples_text += f"\n\n=== 예시 {i} ===\n"
            examples_text += f"키워드: {ex['keyword']}\n"
            examples_text += f"통키워드: {ex['main_keyword_count']}\n"
            examples_text += f"조각키워드: {ex['sub_keyword_count']}\n"
            examples_text += f"서브키워드: {ex['extra_keyword_count']}\n"
            examples_text += f"수정 전:\n{str(ex['original'])[:300]}...\n"
            examples_text += f"수정 후:\n{str(ex['edited'])[:300]}...\n"
        
        prompt = f"""
당신은 블로그 SEO 원고 수정 전문가입니다.

# 핵심 규칙

## 1. 키워드 규칙
- **통 키워드 (핵심 키워드)**: {row_data['main_keyword_count']} 
  → 정확히 이 횟수만큼 반복 (+1까지 허용)
- **조각 키워드**: {row_data['sub_keyword_count']}
  → 각 단어별로 정확히 반복 (+1까지 허용)
- **서브 키워드 목록 수**: {row_data['extra_keyword_count']}개
  → 2회 이상 등장하는 단어의 총 개수 (+1까지 허용)

## 2. 카운팅 규칙 (매우 중요!)
- **띄어쓰기 단위로 카운팅**
- "강남 맛집 추천을" → 통키워드 카운팅 안됨 (조사 '을' 붙음)
- "강남 맛집 추천 리스트" → 통키워드 1회 카운팅 됨
- **한글자 조사(을/를/이/가)**: 띄어쓰기 하지 말고 우회 표현 사용
- **두글자 이상 조사(으로/에게/부터)**: 띄어쓰기 허용

## 3. 첫 문단 필수 규칙
- 핵심 키워드 정확히 2회
- 핵심 키워드 사이에 2문장 이상 삽입

## 4. 글 구조
- **도입부**: 고민/궁금증/경험 소개
- **중간부**: 자연스러운 키워드 반복
- **마무리**: 댓글 유도 (정보 공유 요청, 질문 등)

## 5. 키워드 부족 시
- 자연스러운 문맥에 추가 삽입
- 불가능하면 마지막에 #해시태그 형식으로 추가
- 예: #강남맛집 #맛집추천

## 6. 글자수
- 목표: 약 {row_data['char_count']}자 (±50자 허용)

# 학습 예시 (패턴 참고)
{examples_text}

# 수정할 원고
**키워드**: {row_data['keyword']}

{row_data['original']}

# 지시사항
위 모든 규칙을 정확히 지키면서 자연스럽고 읽기 편한 블로그 글로 수정하세요.
**수정된 원고만 출력**하고, 설명이나 주석은 절대 붙이지 마세요.
"""
        
        return prompt
        
    def process_file(self, input_file):
        """파일 처리"""
        try:
            print("\n" + "="*60)
            print("3️⃣  자동 수정 시작")
            print("="*60)
            
            wb = openpyxl.load_workbook(input_file)
            ws = wb.active
            
            genai.configure(api_key=self.api_key)
            model = genai.GenerativeModel('gemini-2.5-pro')
            
            total_rows = ws.max_row - 1
            
            for row_idx in range(2, ws.max_row + 1):
                print(f"\n{'─'*60}")
                print(f"📄 {row_idx-1}/{total_rows}번째 원고 처리 중...")
                print(f"{'─'*60}")
                
                row_data = {
                    'keyword': ws.cell(row_idx, 2).value,
                    'char_count': ws.cell(row_idx, 3).value,
                    'main_keyword_count': ws.cell(row_idx, 4).value,
                    'sub_keyword_count': ws.cell(row_idx, 5).value,
                    'extra_keyword_count': ws.cell(row_idx, 6).value,
                    'original': ws.cell(row_idx, 7).value
                }
                
                if not row_data['original']:
                    self.log(f"⚠️  {row_idx}행: 원고 없음, 건너뜀")
                    continue
                
                self.log(f"키워드: {row_data['keyword']}")
                self.log(f"목표 글자수: {row_data['char_count']}자")
                
                # 1단계: 기본 교정
                corrected = self.apply_basic_corrections(row_data['original'])
                self.log("✅ 1단계: 기본 교정 완료 (금칙어, 표기법)")
                
                # 2단계: AI 수정
                self.log("⏳ 2단계: AI 수정 중... (10~30초 소요)")
                row_data['original'] = corrected
                prompt = self.create_prompt(row_data)
                
                response = model.generate_content(prompt)
                edited_text = response.text.strip()
                
                # 결과 저장 (H열)
                ws.cell(row_idx, 8).value = edited_text
                self.log(f"✅ 2단계: AI 수정 완료 (결과 글자수: {len(edited_text)}자)")
                
            # 저장
            output_file = input_file.replace('.xlsx', '_수정완료.xlsx')
            wb.save(output_file)
            
            print("\n" + "="*60)
            print("🎉 모든 작업 완료!")
            print("="*60)
            self.log(f"📁 저장 위치: {output_file}")
            
            # 최종 파일 outputs로 복사
            import shutil
            final_output = f"/mnt/user-data/outputs/{os.path.basename(output_file)}"
            shutil.copy(output_file, final_output)
            self.log(f"📥 다운로드 가능: {final_output}")
            
            return output_file
            
        except Exception as e:
            print(f"\n❌ 오류 발생: {str(e)}")
            import traceback
            traceback.print_exc()
            return None
            
def main():
    editor = BlogEditor()
    
    print()
    
    # 금칙어 로딩
    if not editor.load_forbidden_words():
        return
    
    # 예시 로딩
    if not editor.load_examples():
        return
    
    # API 키 입력
    if not editor.input_api_key():
        return
    
    # 파일 선택
    input_file = editor.select_file()
    if not input_file:
        return
    
    # 처리
    result = editor.process_file(input_file)
    
    if result:
        print("\n✅ 프로그램 종료")
    else:
        print("\n❌ 작업 실패")

if __name__ == "__main__":
    main()
