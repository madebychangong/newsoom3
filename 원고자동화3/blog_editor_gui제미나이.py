#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
블로그 원고 자동 수정 프로그램 (GUI 버전)
Gemini API 기반
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import openpyxl
import google.generativeai as genai
import re
import os
from datetime import datetime
import threading
import json
import base64
import random

class BlogEditorGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("블로그 원고 자동 수정 프로그램")
        self.root.geometry("800x700")
        self.root.resizable(False, False)
        
        # 데이터 저장 변수
        self.api_key = ""
        self.forbidden_words = {}
        self.examples = []
        self.input_file = ""
        self.is_processing = False
        self.config_file = os.path.join(os.path.expanduser("~"), ".blog_editor_config.json")

        self.setup_ui()
        self.load_saved_api_key()  # 저장된 API 키 불러오기
        
    def load_saved_api_key(self):
        """저장된 API 키 불러오기"""
        try:
            if os.path.exists(self.config_file):
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    saved_key = config.get('api_key', '')
                    
                    if saved_key:
                        # 간단한 디코딩 (보안은 약하지만 평문보다는 나음)
                        decoded_key = base64.b64decode(saved_key).decode('utf-8')
                        self.api_key = decoded_key
                        self.api_entry.insert(0, decoded_key)
                        self.api_status.config(text="✅ 저장된 API 키 불러옴", fg="green")
                        self.log("✅ 저장된 API 키를 불러왔습니다", "#27ae60")
                        self.check_ready()
        except Exception as e:
            self.log(f"⚠️  저장된 API 키 불러오기 실패: {str(e)}", "#e67e22")
    
    def save_api_key_to_file(self):
        """API 키를 파일로 저장"""
        try:
            # 간단한 인코딩
            encoded_key = base64.b64encode(self.api_key.encode('utf-8')).decode('utf-8')
            
            config = {'api_key': encoded_key}
            
            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(config, f)
                
        except Exception as e:
            self.log(f"⚠️  API 키 저장 실패: {str(e)}", "#e67e22")
        
    def setup_ui(self):
        """UI 구성"""
        # 메인 프레임
        main_frame = ttk.Frame(self.root, padding="15")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 제목
        title_frame = ttk.Frame(main_frame)
        title_frame.pack(fill=tk.X, pady=(0, 20))
        
        title_label = tk.Label(title_frame, text="📝 블로그 원고 자동 수정 프로그램", 
                              font=("맑은 고딕", 18, "bold"), fg="#2c3e50")
        title_label.pack()
        
        subtitle_label = tk.Label(title_frame, text="Gemini 2.5 Pro AI 기반", 
                                 font=("맑은 고딕", 10), fg="#7f8c8d")
        subtitle_label.pack()
        
        # 1. API 키 입력 섹션
        api_frame = ttk.LabelFrame(main_frame, text="  1️⃣  Gemini API 키 입력  ", padding="10")
        api_frame.pack(fill=tk.X, pady=(0, 10))
        
        api_input_frame = ttk.Frame(api_frame)
        api_input_frame.pack(fill=tk.X)
        
        self.api_entry = ttk.Entry(api_input_frame, width=50, show="*", font=("맑은 고딕", 10))
        self.api_entry.pack(side=tk.LEFT, padx=(0, 10), fill=tk.X, expand=True)
        
        self.api_button = ttk.Button(api_input_frame, text="저장", command=self.save_api_key)
        self.api_button.pack(side=tk.LEFT)
        
        self.api_status = tk.Label(api_frame, text="❌ API 키 미등록", 
                                   font=("맑은 고딕", 9), fg="red")
        self.api_status.pack(anchor=tk.W, pady=(5, 0))
        
        api_help = tk.Label(api_frame, text="💡 API 키 발급: https://aistudio.google.com/app/apikey", 
                           font=("맑은 고딕", 8), fg="#3498db", cursor="hand2")
        api_help.pack(anchor=tk.W)
        api_help.bind("<Button-1>", lambda e: self.open_url("https://aistudio.google.com/app/apikey"))
        
        # 2. 파일 선택 섹션
        file_frame = ttk.LabelFrame(main_frame, text="  2️⃣  수정할 엑셀 파일 선택  ", padding="10")
        file_frame.pack(fill=tk.X, pady=(0, 10))
        
        file_input_frame = ttk.Frame(file_frame)
        file_input_frame.pack(fill=tk.X)
        
        self.file_label = tk.Label(file_input_frame, text="선택된 파일 없음", 
                                   font=("맑은 고딕", 9), fg="gray", anchor=tk.W)
        self.file_label.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        self.file_button = ttk.Button(file_input_frame, text="📁 파일 선택", command=self.select_file)
        self.file_button.pack(side=tk.LEFT)
        
        file_help = tk.Label(file_frame, text="💡 같은 폴더에 금칙어_리스트.xlsx, 수정전후.xlsx, 블로그_작업_엑셀템플릿.xlsx 필요", 
                            font=("맑은 고딕", 8), fg="#7f8c8d")
        file_help.pack(anchor=tk.W, pady=(5, 0))
        
        # 3. 실행 버튼
        run_frame = ttk.LabelFrame(main_frame, text="  3️⃣  자동 수정 실행  ", padding="10")
        run_frame.pack(fill=tk.X, pady=(0, 10))
        
        self.run_button = ttk.Button(run_frame, text="🚀 자동 수정 시작", 
                                     command=self.start_processing, state='disabled')
        self.run_button.pack(fill=tk.X)
        
        # 4. 진행 상황
        progress_frame = ttk.LabelFrame(main_frame, text="  처리 상황  ", padding="10")
        progress_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        self.progress_text = scrolledtext.ScrolledText(progress_frame, width=80, height=20, 
                                                       wrap=tk.WORD, font=("맑은 고딕", 9))
        self.progress_text.pack(fill=tk.BOTH, expand=True)
        
        # 하단 상태바
        status_frame = ttk.Frame(main_frame)
        status_frame.pack(fill=tk.X)
        
        self.status_label = tk.Label(status_frame, text="대기 중...", 
                                     font=("맑은 고딕", 9), fg="#3498db", anchor=tk.W)
        self.status_label.pack(side=tk.LEFT)
        
    def open_url(self, url):
        """URL 열기"""
        import webbrowser
        webbrowser.open(url)
        
    def log(self, message, color=None):
        """로그 출력"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.progress_text.insert(tk.END, f"[{timestamp}] {message}\n")
        if color:
            # 마지막 줄에 색상 적용
            line_start = self.progress_text.index("end-2c linestart")
            line_end = self.progress_text.index("end-1c")
            self.progress_text.tag_add(color, line_start, line_end)
            self.progress_text.tag_config(color, foreground=color)
        self.progress_text.see(tk.END)
        self.root.update()
        
    def save_api_key(self):
        """API 키 저장 (검증 없이)"""
        self.api_key = self.api_entry.get().strip()
        
        if not self.api_key:
            messagebox.showwarning("경고", "API 키를 입력해주세요.")
            return
        
        # 파일로 저장
        self.save_api_key_to_file()
        
        self.api_status.config(text="✅ API 키 저장 완료 (gemini-2.5-pro)", fg="green")
        self.log("✅ API 키가 저장되었습니다", "#27ae60")
        self.check_ready()
        messagebox.showinfo("저장 완료", "API 키가 저장되었습니다.\n다음 실행 시 자동으로 불러옵니다.")
            
    def select_file(self):
        """파일 선택"""
        filename = filedialog.askopenfilename(
            title="수정할 엑셀 파일 선택",
            filetypes=[("Excel 파일", "*.xlsx"), ("모든 파일", "*.*")]
        )
        
        if filename:
            self.input_file = filename
            basename = os.path.basename(filename)
            self.file_label.config(text=f"📄 {basename}", fg="green")
            self.log(f"파일 선택: {basename}", "#3498db")
            self.check_ready()
            
    def check_ready(self):
        """실행 가능 여부 체크"""
        if self.api_key and self.input_file:
            self.run_button.config(state='normal')
            self.status_label.config(text="✅ 준비 완료 - 실행 버튼을 눌러주세요", fg="green")
        else:
            self.run_button.config(state='disabled')
            
    def start_processing(self):
        """처리 시작 (별도 스레드)"""
        if self.is_processing:
            return
        
        self.is_processing = True
        self.run_button.config(state='disabled')
        self.file_button.config(state='disabled')
        self.api_button.config(state='disabled')
        
        # 별도 스레드에서 처리
        thread = threading.Thread(target=self.process_file)
        thread.daemon = True
        thread.start()
        
    def load_forbidden_words(self, base_dir):
        """금칙어 로딩"""
        try:
            file_path = os.path.join(base_dir, '금칙어_리스트.xlsx')
            
            if not os.path.exists(file_path):
                self.log(f"⚠️  금칙어 파일 없음: {file_path}", "#e67e22")
                return False
            
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            for row_idx in range(3, ws.max_row + 1):
                forbidden = ws.cell(row_idx, 2).value
                alternatives = []
                
                for col_idx in range(3, 10):
                    alt = ws.cell(row_idx, col_idx).value
                    if alt:
                        alternatives.append(str(alt).strip())
                
                if forbidden and alternatives:
                    self.forbidden_words[str(forbidden).strip()] = alternatives
            
            self.log(f"✅ 금칙어 {len(self.forbidden_words)}개 로딩 완료", "#27ae60")
            return True
            
        except Exception as e:
            self.log(f"❌ 금칙어 로딩 실패: {str(e)}", "#e74c3c")
            return False
            
    def load_examples(self, base_dir):
        """학습 예시 로딩"""
        try:
            # 수정전후.xlsx
            file1 = os.path.join(base_dir, '수정전후.xlsx')
            if os.path.exists(file1):
                wb1 = openpyxl.load_workbook(file1)
                ws1 = wb1.active
                
                for row_idx in range(2, ws1.max_row + 1):
                    example = {
                        'keyword': ws1.cell(row_idx, 2).value,
                        'char_count': ws1.cell(row_idx, 3).value,
                        'main_keyword_count': ws1.cell(row_idx, 4).value,
                        'sub_keyword_count': ws1.cell(row_idx, 5).value,
                        'extra_keyword_count': ws1.cell(row_idx, 6).value,
                        'original': ws1.cell(row_idx, 7).value,
                        'edited': ws1.cell(row_idx, 8).value
                    }
                    if example['original'] and example['edited']:
                        self.examples.append(example)
            
            # 블로그_작업_엑셀템플릿.xlsx
            file2 = os.path.join(base_dir, '블로그_작업_엑셀템플릿.xlsx')
            if os.path.exists(file2):
                wb2 = openpyxl.load_workbook(file2)
                ws2 = wb2.active
                
                for row_idx in range(2, ws2.max_row + 1):
                    example = {
                        'keyword': ws2.cell(row_idx, 2).value,
                        'char_count': ws2.cell(row_idx, 3).value,
                        'main_keyword_count': ws2.cell(row_idx, 4).value,
                        'sub_keyword_count': ws2.cell(row_idx, 5).value,
                        'extra_keyword_count': ws2.cell(row_idx, 6).value,
                        'original': ws2.cell(row_idx, 7).value,
                        'edited': ws2.cell(row_idx, 8).value
                    }
                    if example['original'] and example['edited']:
                        self.examples.append(example)
            
            self.log(f"✅ 학습 예시 {len(self.examples)}개 로딩 완료", "#27ae60")
            return len(self.examples) > 0
            
        except Exception as e:
            self.log(f"❌ 예시 로딩 실패: {str(e)}", "#e74c3c")
            return False
            
    def analyze_speaker(self, text, model):
        """화자 정보 분석 (성별, 연령대, 상황)"""
        if not text:
            return "분석 불가"
        
        try:
            analysis_prompt = f"""
다음 블로그 글을 분석하여 작성자(화자)의 정보를 유추해주세요.

글:
{text[:500]}...

다음 형식으로만 답변하세요 (다른 설명 없이):
성별: [남성/여성/알 수 없음]
연령대: [20대/30대/40대/50대/60대 이상/알 수 없음]
상황: [한 줄로 간단히 설명]

예시:
성별: 여성
연령대: 30대
상황: 자녀 키 성장 고민
"""
            
            response = model.generate_content(analysis_prompt)
            analysis = response.text.strip()
            
            # 한 줄로 정리
            analysis = analysis.replace('\n', ' / ')
            
            return analysis
            
        except Exception as e:
            return f"분석 실패: {str(e)}"
    
    def add_line_breaks(self, text):
        """자동 문단 구분 추가 (첫 문단 5문장, 이후 3-4문장 랜덤)"""
        if not text:
            return text

        # 기존 줄바꿈 정리
        text = text.replace('\n', ' ').strip()

        # 문장 분리 (., !, ? 기준)
        sentences = re.split(r'([.!?])\s+', text)

        # 문장 재조립 (종결부호 포함)
        full_sentences = []
        for i in range(0, len(sentences)-1, 2):
            if i+1 < len(sentences):
                full_sentences.append(sentences[i] + sentences[i+1])

        # 마지막 문장 처리
        if len(sentences) % 2 == 1:
            full_sentences.append(sentences[-1])

        if not full_sentences:
            return text

        # 문단 구성
        paragraphs = []
        idx = 0

        # 첫 문단: 5문장
        if len(full_sentences) >= 5:
            first_para = ' '.join(full_sentences[:5])
            paragraphs.append(first_para)
            idx = 5
        else:
            # 전체 문장이 5개 미만이면 그대로
            return text

        # 나머지 문단: 3-4문장씩 랜덤
        while idx < len(full_sentences):
            para_size = random.choice([3, 4])
            para_sentences = full_sentences[idx:idx+para_size]
            if para_sentences:
                paragraphs.append(' '.join(para_sentences))
            idx += para_size

        # 문단 사이에 빈 줄 추가
        result = '\n\n'.join(paragraphs)

        return result.strip()
    
    def apply_basic_corrections(self, text):
        """기본 교정"""
        if not text:
            return text
        
        # 1. 네요 -> 내요 (무조건)
        text = text.replace('네요', '내요')
        
        # 2. 더라 -> 더 라 (무조건)
        text = text.replace('더라', '더 라')
        
        # 3. 이모티콘 앞뒤 띄어쓰기 처리 (서브키워드 카운팅을 위해)
        emoticons = ['^^', '??', '!!', '~~', '...', 'ㅠㅠ', 'TT', 'ㅎㅎ', ';;', '--', 'ㅋㅋ']
        
        for emoticon in emoticons:
            # 이모티콘 앞에 띄어쓰기 없으면 추가
            # "좋아요^^" → "좋아요 ^^"
            text = re.sub(r'([^\s])' + re.escape(emoticon), r'\1 ' + emoticon, text)
            
            # 이모티콘 뒤 문장부호 제거하고 띄어쓰기
            # "^^ ." → "^^ "
            text = text.replace(f'{emoticon}.', f'{emoticon} ')
            text = text.replace(f'{emoticon},', f'{emoticon} ')
            text = text.replace(f'{emoticon}!', f'{emoticon} ')
            text = text.replace(f'{emoticon}?', f'{emoticon} ')
            
            # 이모티콘 뒤에 아무것도 없거나 문자가 바로 오면 띄어쓰기 추가
            # "^^ 다음" 은 그대로, "^^다음" → "^^ 다음"
            text = re.sub(re.escape(emoticon) + r'([^\s.,!?])', emoticon + r' \1', text)
        
        # 4. 금칙어 치환
        for forbidden, alternatives in self.forbidden_words.items():
            if forbidden in text and alternatives:
                text = text.replace(forbidden, alternatives[0])
        
        return text
    
    def clean_markdown(self, text):
        """마크다운 형식 제거"""
        if not text:
            return text
        
        # ** 강조 제거
        text = re.sub(r'\*\*([^*]+)\*\*', r'\1', text)
        
        # * 강조 제거
        text = re.sub(r'\*([^*]+)\*', r'\1', text)
        
        # # 헤더 제거
        text = re.sub(r'^#+\s+', '', text, flags=re.MULTILINE)
        
        # 마크다운 코드 블록 제거 (```로 둘러싸인 부분)
        text = re.sub(r'```[^`]*```', '', text, flags=re.DOTALL)
        
        return text.strip()
    
    def parse_keyword_rule(self, rule_text):
        """키워드 규칙 파싱"""
        if not rule_text:
            return ""
        
        rule_text = str(rule_text).strip()
        
        # "키워드 : 숫자" 형식 파싱
        match = re.match(r'(.+?)\s*:\s*(\d+)', rule_text)
        if match:
            keyword = match.group(1).strip()
            count = match.group(2).strip()
            return f"'{keyword}'를 정확히 {count}번 반복 (±1 허용)"
        
        return rule_text
    
    def parse_sub_keywords(self, rule_text):
        """조각 키워드 규칙 파싱"""
        if not rule_text:
            return ""
        
        rule_text = str(rule_text).strip()
        
        # 여러 줄로 나뉜 경우 처리
        lines = rule_text.split('\n')
        parsed_rules = []
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
            
            # "키워드 : 숫자" 형식 파싱
            match = re.match(r'(.+?)\s*:\s*(\d+)', line)
            if match:
                keyword = match.group(1).strip()
                count = match.group(2).strip()
                parsed_rules.append(f"'{keyword}' {count}번")
        
        if parsed_rules:
            return ", ".join(parsed_rules) + " 각각 반복 (±1 허용)"
        
        return rule_text
        
    def create_stage1_prompt(self, row_data, original_text):
        """1단계 프롬프트: 글자수와 키워드 자연스럽게 삽입"""

        # 키워드 규칙 파싱
        main_keyword_rule = self.parse_keyword_rule(row_data['main_keyword_count'])
        sub_keyword_rule = self.parse_sub_keywords(row_data['sub_keyword_count'])

        # 글자수 및 오차 계산
        target_chars = int(row_data['char_count']) if row_data['char_count'] else 1000
        char_tolerance = int(target_chars * 0.05)

        # 금칙어 리스트 생성
        forbidden_list = ""
        for forbidden, alternatives in self.forbidden_words.items():
            alt_text = ", ".join(alternatives[:3])
            forbidden_list += f"- '{forbidden}' 대신 → {alt_text}\n"

        prompt = f"""당신은 원고 수정 전문가입니다. [1단계 작업]

# 목표
기존 원고를 최대한 보존하며 글자수와 키워드를 자연스럽게 조정하세요.

# 규칙 (간단함!)
1. 글자수: {target_chars - char_tolerance}~{target_chars + char_tolerance}자 (최우선!)
2. 핵심키워드 '{row_data['keyword']}': 첫문단 약 2회 + 나머지 {main_keyword_rule} (대략적으로)
3. 조각키워드: {sub_keyword_rule} (대략적으로)
4. 금칙어 대체:
{forbidden_list}
5. 원문 톤/스타일 유지

# 수정할 원고
{original_text}

# 지시
수정된 원고만 출력하세요. 설명 금지.
"""
        return prompt

    def create_stage2_prompt(self, row_data, stage1_result):
        """2단계 프롬프트: 세부 규칙 정확히 적용"""

        extra_keyword_count = str(row_data['extra_keyword_count']).strip() if row_data['extra_keyword_count'] else "0"
        keyword_start_count = str(row_data['keyword_start_count']).strip() if row_data['keyword_start_count'] else "2~3"

        prompt = f"""당신은 원고 정제 전문가입니다. [2단계 작업]

⚠️ 중요: 글자수를 절대 변경하지 마세요! 1단계에서 이미 맞췄습니다.
띄어쓰기와 구조만 조정하세요.

# 1단계 결과
{stage1_result}

# 세부 규칙
1. 첫 문단 구조
   - 정확히 4-5문장
   - 핵심키워드 '{row_data['keyword']}' 정확히 2회
   - 키워드 사이에 최소 2문장 삽입

2. 키워드 띄어쓰기 (절대 규칙!)
   - 모든 키워드 뒤 띄어쓰기 필수
   - "추천을" (X) → "추천 정보를" (O)
   - 한글자 조사(을/를/이/가) 절대 금지!
   - 우회 표현 사용: 정보/내용/방법/리스트/관련/사항

3. 핵심키워드 시작 문장: {keyword_start_count}개

4. 서브키워드: {extra_keyword_count}개 (부족시 ^^, ??, .. 활용)

# 지시
수정된 원고만 출력하세요. 설명 금지. 글자수 변경 절대 금지!
"""
        return prompt
        
    def process_file(self):
        """파일 처리 메인 로직"""
        try:
            self.log("\n" + "="*60, "#2c3e50")
            self.log("🚀 자동 수정 시작...", "#2c3e50")
            self.log("="*60, "#2c3e50")
            self.status_label.config(text="⏳ 처리 중...", fg="orange")
            
            # 같은 폴더 경로
            base_dir = os.path.dirname(self.input_file)
            
            # 금칙어 로딩
            if not self.load_forbidden_words(base_dir):
                messagebox.showwarning("경고", "금칙어 파일을 찾을 수 없습니다.\n같은 폴더에 '금칙어_리스트.xlsx'를 넣어주세요.")
            
            # 예시 로딩
            if not self.load_examples(base_dir):
                messagebox.showwarning("경고", "예시 파일을 찾을 수 없습니다.\n같은 폴더에 '수정전후.xlsx', '블로그_작업_엑셀템플릿.xlsx'를 넣어주세요.")
            
            # 입력 파일 로드
            wb = openpyxl.load_workbook(self.input_file)
            ws = wb.active
            
            # Gemini 모델 초기화
            genai.configure(api_key=self.api_key)
            model = genai.GenerativeModel('gemini-2.5-pro')
            
            total_rows = ws.max_row - 1
            
            for row_idx in range(2, ws.max_row + 1):
                self.log(f"\n{'─'*60}", "#95a5a6")
                self.log(f"📄 {row_idx-1}/{total_rows}번째 원고 처리 중...", "#3498db")
                self.log(f"{'─'*60}", "#95a5a6")
                
                # 데이터 추출 (새 양식)
                row_data = {
                    'keyword': ws.cell(row_idx, 2).value,  # B열: 키워드
                    'main_keyword_count': ws.cell(row_idx, 4).value,  # D열: 통키워드 반복수
                    'sub_keyword_count': ws.cell(row_idx, 5).value,  # E열: 조각키워드 반복수
                    'original': ws.cell(row_idx, 7).value,  # G열: 원고
                    'char_count': ws.cell(row_idx, 10).value,  # J열: 실제 글자수
                    'keyword_start_count': ws.cell(row_idx, 11).value,  # K열: 문장시작통키워드 수
                    'extra_keyword_count': ws.cell(row_idx, 12).value  # L열: 보정 서브키워드 목록 수
                }
                
                if not row_data['original']:
                    self.log(f"⚠️  {row_idx}행: 원고 없음, 건너뜀", "#e67e22")
                    continue
                
                self.log(f"키워드: {row_data['keyword']}")
                self.log(f"목표 글자수: {row_data['char_count']}자")

                # 원고에서 제목 제거 (# 으로 시작하는 첫 줄)
                original_text = str(row_data['original'])
                lines = original_text.split('\n')
                if lines and lines[0].strip().startswith('#'):
                    original_text = '\n'.join(lines[1:]).strip()
                    self.log("📌 제목 라인 제거됨", "#95a5a6")

                # [1단계] AI 수정 - 글자수 + 키워드 자연스럽게
                self.log("⏳ [1단계] 글자수 및 키워드 조정 중...", "#f39c12")
                stage1_prompt = self.create_stage1_prompt(row_data, original_text)

                response1 = model.generate_content(stage1_prompt)
                stage1_result = response1.text.strip()
                stage1_result = self.clean_markdown(stage1_result)

                self.log(f"✅ [1단계] 완료 (글자수: {len(stage1_result)}자)", "#27ae60")

                # [2단계] AI 수정 - 세부 규칙 정확히 적용
                self.log("⏳ [2단계] 세부 규칙 적용 중...", "#f39c12")
                stage2_prompt = self.create_stage2_prompt(row_data, stage1_result)

                response2 = model.generate_content(stage2_prompt)
                edited_text = response2.text.strip()
                edited_text = self.clean_markdown(edited_text)

                self.log(f"✅ [2단계] 완료 (글자수: {len(edited_text)}자)", "#27ae60")

                # AI 생성 후 기본 교정 적용 (네요→내요, 더라→더 라, 금칙어)
                edited_text = self.apply_basic_corrections(edited_text)

                # 자동 문단 구분 추가 (첫 5문장, 이후 3-4문장)
                edited_text = self.add_line_breaks(edited_text)

                # 결과 저장 (M열 = 13번)
                ws.cell(row_idx, 13).value = edited_text
                self.log(f"✅ 최종 완료 (결과 글자수: {len(edited_text)}자)", "#27ae60")
                
                # 화자 분석 (N열 = 14번)
                self.log("⏳ 화자 정보 분석 중...", "#3498db")
                speaker_info = self.analyze_speaker(edited_text, model)
                ws.cell(row_idx, 14).value = speaker_info
                self.log(f"✅ 화자 분석 완료: {speaker_info}", "#27ae60")
                
            # 결과 파일 저장 (원본 파일에 덮어쓰기)
            wb.save(self.input_file)
            
            self.log("\n" + "="*60, "#2c3e50")
            self.log("🎉 모든 작업 완료!", "#27ae60")
            self.log("="*60, "#2c3e50")
            self.log(f"📁 저장 위치: {self.input_file}", "#3498db")
            
            self.status_label.config(text="✅ 완료!", fg="green")
            
            messagebox.showinfo("완료", f"수정이 완료되었습니다!\n\n원본 파일에 저장됨:\n{self.input_file}")
            
        except Exception as e:
            self.log(f"\n❌ 오류 발생: {str(e)}", "#e74c3c")
            self.status_label.config(text="❌ 오류 발생", fg="red")
            messagebox.showerror("오류", f"처리 중 오류가 발생했습니다:\n{str(e)}")
            
        finally:
            self.is_processing = False
            self.run_button.config(state='normal')
            self.file_button.config(state='normal')
            self.api_button.config(state='normal')

def main():
    root = tk.Tk()
    app = BlogEditorGUI(root)
    root.mainloop()

if __name__ == "__main__":
    main()